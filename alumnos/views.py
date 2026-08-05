import re

from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render

from openpyxl import load_workbook

from .forms import AlumnoForm, ImportarAlumnosForm
from .models import Alumno


def alumno_lista(request):
    busqueda = request.GET.get('buscar', '').strip()

    alumnos = Alumno.objects.all()

    if busqueda:
        alumnos = alumnos.filter(
            Q(cedula__icontains=busqueda)
            | Q(nombre_completo__icontains=busqueda)
            | Q(nombres__icontains=busqueda)
            | Q(apellidos__icontains=busqueda)
            | Q(curso__icontains=busqueda)
            | Q(especialidad__icontains=busqueda)
            | Q(seccion__icontains=busqueda)
        )

    contexto = {
        'alumnos': alumnos,
        'busqueda': busqueda,
    }

    return render(
        request,
        'alumnos/alumno_lista.html',
        contexto,
    )


def alumno_crear(request):
    if request.method == 'POST':
        formulario = AlumnoForm(request.POST)

        if formulario.is_valid():
            alumno = formulario.save()

            messages.success(
                request,
                (
                    f'El alumno {alumno.nombre_visible} '
                    'fue registrado correctamente.'
                ),
            )

            return redirect('alumnos:lista')
    else:
        formulario = AlumnoForm()

    contexto = {
        'formulario': formulario,
        'titulo': 'Registrar alumno',
        'texto_boton': 'Guardar alumno',
    }

    return render(
        request,
        'alumnos/alumno_formulario.html',
        contexto,
    )


def alumno_editar(request, alumno_id):
    alumno = get_object_or_404(
        Alumno,
        id=alumno_id,
    )

    if request.method == 'POST':
        formulario = AlumnoForm(
            request.POST,
            instance=alumno,
        )

        if formulario.is_valid():
            formulario.save()

            messages.success(
                request,
                'Los datos del alumno fueron actualizados.',
            )

            return redirect('alumnos:lista')
    else:
        formulario = AlumnoForm(instance=alumno)

    contexto = {
        'formulario': formulario,
        'titulo': 'Editar alumno',
        'texto_boton': 'Actualizar alumno',
        'alumno': alumno,
    }

    return render(
        request,
        'alumnos/alumno_formulario.html',
        contexto,
    )


def limpiar_texto(valor):
    if valor is None:
        return ''

    return ' '.join(str(valor).strip().split())


def limpiar_cedula(valor):
    if valor is None:
        return ''

    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)

    cedula = str(valor).strip()

    cedula = (
        cedula
        .replace('.', '')
        .replace(' ', '')
        .replace('-', '')
    )

    return cedula


def obtener_datos_grupo(texto):
    """
    Obtiene el curso, la especialidad y la sección desde textos como:

    CURSO: 1° Electricidad Sección "A"
    CURSO: 2° Electrónica Sección "B"
    """
    texto = limpiar_texto(texto).upper()

    patron = (
        r'(\d)\s*°?.*?'
        r'(ELECTRICIDAD|ELECTR[ÓO]NICA).*?'
        r'SECCI[ÓO]N\s*"?([AB])"?'
    )

    coincidencia = re.search(patron, texto)

    if not coincidencia:
        return None

    numero_curso = coincidencia.group(1)
    especialidad = coincidencia.group(2)
    seccion = coincidencia.group(3)

    if especialidad == 'ELECTRONICA':
        especialidad = 'ELECTRÓNICA'

    return {
        'curso': f'{numero_curso}°',
        'especialidad': especialidad,
        'seccion': seccion,
    }


def obtener_grupos_excel(hoja):
    """
    Localiza los doce grupos colocados horizontalmente.

    Cada grupo utiliza tres columnas:
    número, nombre completo y cédula.
    """
    columnas_iniciales = [1, 7, 13, 19]
    grupos = []

    for numero_fila in range(1, hoja.max_row + 1):
        for columna_inicial in columnas_iniciales:
            encabezado = hoja.cell(
                row=numero_fila,
                column=columna_inicial,
            ).value

            datos_grupo = obtener_datos_grupo(encabezado)

            if datos_grupo:
                grupos.append({
                    'fila_encabezado': numero_fila,
                    'columna_numero': columna_inicial,
                    'columna_nombre': columna_inicial + 1,
                    'columna_cedula': columna_inicial + 2,
                    **datos_grupo,
                })

    return grupos


def obtener_fila_final_grupo(grupo, grupos, hoja):
    """
    Determina dónde termina cada bloque de alumnos.
    """
    siguientes_filas = [
        otro_grupo['fila_encabezado']
        for otro_grupo in grupos
        if otro_grupo['fila_encabezado']
        > grupo['fila_encabezado']
    ]

    if siguientes_filas:
        return min(siguientes_filas) - 1

    return hoja.max_row


def alumno_importar(request):
    resultados = None

    if request.method == 'POST':
        formulario = ImportarAlumnosForm(
            request.POST,
            request.FILES,
        )

        if formulario.is_valid():
            archivo = formulario.cleaned_data['archivo']

            try:
                libro_excel = load_workbook(
                    archivo,
                    read_only=True,
                    data_only=True,
                )

                if 'Eldad' not in libro_excel.sheetnames:
                    formulario.add_error(
                        'archivo',
                        (
                            'El archivo debe contener una hoja '
                            'llamada Eldad.'
                        ),
                    )
                else:
                    hoja = libro_excel['Eldad']
                    grupos = obtener_grupos_excel(hoja)

                    if len(grupos) != 12:
                        formulario.add_error(
                            'archivo',
                            (
                                'No se pudieron reconocer los 12 grupos. '
                                f'Se encontraron {len(grupos)}.'
                            ),
                        )
                    else:
                        resultados = procesar_alumnos_excel(
                            hoja,
                            grupos,
                        )

                libro_excel.close()

            except Exception as error:
                formulario.add_error(
                    'archivo',
                    (
                        'No se pudo leer la planilla. '
                        f'Detalle: {error}'
                    ),
                )
    else:
        formulario = ImportarAlumnosForm()

    contexto = {
        'formulario': formulario,
        'resultados': resultados,
    }

    return render(
        request,
        'alumnos/alumno_importar.html',
        contexto,
    )


@transaction.atomic
def procesar_alumnos_excel(hoja, grupos):
    creados = 0
    actualizados = 0
    repetidos_planilla = 0
    errores = []
    cedulas_planilla = set()

    for grupo in grupos:
        fila_inicial = grupo['fila_encabezado'] + 2

        fila_final = obtener_fila_final_grupo(
            grupo,
            grupos,
            hoja,
        )

        for numero_fila in range(
            fila_inicial,
            fila_final + 1,
        ):
            nombre_completo = limpiar_texto(
                hoja.cell(
                    row=numero_fila,
                    column=grupo['columna_nombre'],
                ).value
            )

            cedula = limpiar_cedula(
                hoja.cell(
                    row=numero_fila,
                    column=grupo['columna_cedula'],
                ).value
            )

            # Las filas completamente vacías no se procesan.
            if not nombre_completo and not cedula:
                continue

            if not nombre_completo:
                errores.append(
                    (
                        f'Fila {numero_fila}: falta el nombre '
                        'del alumno.'
                    )
                )
                continue

            if not cedula:
                errores.append(
                    f'Fila {numero_fila}: falta la cédula.'
                )
                continue

            # Admite cédulas nacionales y extranjeras.
            if not cedula.isdigit():
                errores.append(
                    (
                        f'Fila {numero_fila}: la cédula {cedula} '
                        'debe contener solamente números.'
                    )
                )
                continue

            if len(cedula) > 20:
                errores.append(
                    (
                        f'Fila {numero_fila}: la cédula {cedula} '
                        'es demasiado extensa.'
                    )
                )
                continue

            if cedula in cedulas_planilla:
                repetidos_planilla += 1

                errores.append(
                    (
                        f'Fila {numero_fila}: la cédula {cedula} '
                        'está repetida en la planilla.'
                    )
                )
                continue

            cedulas_planilla.add(cedula)

            alumno, creado = Alumno.objects.update_or_create(
                cedula=cedula,
                defaults={
                    'nombre_completo': nombre_completo,
                    'curso': grupo['curso'],
                    'especialidad': grupo['especialidad'],
                    'seccion': grupo['seccion'],
                    'turno': 'MAÑANA',
                    'activo': True,
                },
            )

            if creado:
                creados += 1
            else:
                actualizados += 1

    return {
        # Se conserva "importados" para que funcione la plantilla actual.
        'importados': creados,
        'creados': creados,
        'actualizados': actualizados,
        'duplicados': repetidos_planilla,
        'errores': errores,
        'cantidad_errores': len(errores),
        'total_procesados': creados + actualizados,
    }