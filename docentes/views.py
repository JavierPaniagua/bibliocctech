from django.contrib import messages
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import load_workbook

from .forms import DocenteForm, ImportarDocentesForm
from .models import Docente


def docente_lista(request):
    busqueda = request.GET.get('buscar', '').strip()

    docentes = Docente.objects.all()

    if busqueda:
        docentes = docentes.filter(
            Q(cedula__icontains=busqueda)
            | Q(nombres__icontains=busqueda)
            | Q(apellidos__icontains=busqueda)
            | Q(area__icontains=busqueda)
        )

    contexto = {
        'docentes': docentes,
        'busqueda': busqueda,
    }

    return render(
        request,
        'docentes/docente_lista.html',
        contexto,
    )


def docente_crear(request):
    if request.method == 'POST':
        formulario = DocenteForm(request.POST)

        if formulario.is_valid():
            docente = formulario.save()

            messages.success(
                request,
                (
                    f'El docente {docente.nombres} '
                    f'{docente.apellidos} fue registrado correctamente.'
                ),
            )

            return redirect('docentes:lista')
    else:
        formulario = DocenteForm()

    contexto = {
        'formulario': formulario,
        'titulo': 'Registrar docente',
        'texto_boton': 'Guardar docente',
    }

    return render(
        request,
        'docentes/docente_formulario.html',
        contexto,
    )


def docente_editar(request, docente_id):
    docente = get_object_or_404(
        Docente,
        id=docente_id,
    )

    if request.method == 'POST':
        formulario = DocenteForm(
            request.POST,
            instance=docente,
        )

        if formulario.is_valid():
            formulario.save()

            messages.success(
                request,
                'Los datos del docente fueron actualizados.',
            )

            return redirect('docentes:lista')
    else:
        formulario = DocenteForm(instance=docente)

    contexto = {
        'formulario': formulario,
        'titulo': 'Editar docente',
        'texto_boton': 'Actualizar docente',
        'docente': docente,
    }

    return render(
        request,
        'docentes/docente_formulario.html',
        contexto,
    )

def limpiar_texto(valor):
    if valor is None:
        return ''

    return str(valor).strip()


def limpiar_cedula(valor):
    if valor is None:
        return ''

    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)

    cedula = str(valor).strip()

    cedula = (
        cedula
        .replace('.', '')
        .replace(' ', '')
        .replace('-', '')
    )

    return cedula


def convertir_activo(valor):
    texto = limpiar_texto(valor).upper()

    return texto in {
        'SI',
        'SÍ',
        'TRUE',
        'VERDADERO',
        '1',
        'ACTIVO',
    }


def docente_importar(request):
    resultados = None

    if request.method == 'POST':
        formulario = ImportarDocentesForm(
            request.POST,
            request.FILES,
        )

        if formulario.is_valid():
            archivo = formulario.cleaned_data['archivo']

            try:
                libro_excel = load_workbook(
                    archivo,
                    read_only=True,
                    data_only=True,
                )

                if 'Docentes' not in libro_excel.sheetnames:
                    formulario.add_error(
                        'archivo',
                        (
                            'El archivo debe contener una hoja '
                            'llamada Docentes.'
                        ),
                    )
                else:
                    hoja = libro_excel['Docentes']

                    encabezados_esperados = [
                        'cedula',
                        'nombres',
                        'apellidos',
                        'area',
                        'especialidad',
                        'telefono',
                        'correo',
                        'turno',
                        'activo',
                    ]

                    encabezados_archivo = [
                        limpiar_texto(
                            hoja.cell(
                                row=4,
                                column=columna,
                            ).value
                        ).lower()
                        for columna in range(1, 10)
                    ]

                    if encabezados_archivo != encabezados_esperados:
                        formulario.add_error(
                            'archivo',
                            (
                                'Las columnas fueron modificadas. '
                                'Utilice la plantilla oficial.'
                            ),
                        )
                    else:
                        resultados = procesar_docentes_excel(hoja)

            except Exception:
                formulario.add_error(
                    'archivo',
                    (
                        'No se pudo leer la planilla. Compruebe '
                        'que sea un archivo Excel válido.'
                    ),
                )
    else:
        formulario = ImportarDocentesForm()

    contexto = {
        'formulario': formulario,
        'resultados': resultados,
    }

    return render(
        request,
        'docentes/docente_importar.html',
        contexto,
    )


def procesar_docentes_excel(hoja):
    especialidades_validas = {
        'ELECTRÓNICA',
        'ELECTRONICA',
        'ELECTRICIDAD',
    }

    turnos_validos = {
        'MAÑANA',
        'TARDE',
        'NOCHE',
    }

    importados = 0
    duplicados = 0
    errores = []
    cedulas_planilla = set()

    for numero_fila, fila in enumerate(
        hoja.iter_rows(
            min_row=5,
            max_col=9,
            values_only=True,
        ),
        start=5,
    ):
        if all(valor is None for valor in fila):
            continue

        (
            cedula,
            nombres,
            apellidos,
            area,
            especialidad,
            telefono,
            correo,
            turno,
            activo,
        ) = fila

        cedula = limpiar_cedula(cedula)
        nombres = limpiar_texto(nombres)
        apellidos = limpiar_texto(apellidos)
        area = limpiar_texto(area)
        especialidad = limpiar_texto(especialidad).upper()
        telefono = limpiar_texto(telefono)
        correo = limpiar_texto(correo).lower()
        turno = limpiar_texto(turno).upper()

        if not cedula or not nombres or not apellidos:
            errores.append(
                f'Fila {numero_fila}: faltan datos obligatorios.'
            )
            continue

        if not cedula.isdigit():
            errores.append(
                f'Fila {numero_fila}: la cédula debe contener '
                f'solamente números.'
            )
            continue

        if especialidad not in especialidades_validas:
            errores.append(
                f'Fila {numero_fila}: especialidad incorrecta.'
            )
            continue

        if especialidad == 'ELECTRONICA':
            especialidad = 'ELECTRÓNICA'

        if turno not in turnos_validos:
            errores.append(
                f'Fila {numero_fila}: turno incorrecto.'
            )
            continue

        if correo:
            try:
                validate_email(correo)
            except ValidationError:
                errores.append(
                    f'Fila {numero_fila}: correo electrónico '
                    f'incorrecto.'
                )
                continue

        if cedula in cedulas_planilla:
            duplicados += 1

            errores.append(
                f'Fila {numero_fila}: la cédula {cedula} '
                f'está repetida en la planilla.'
            )

            continue

        cedulas_planilla.add(cedula)

        if Docente.objects.filter(cedula=cedula).exists():
            duplicados += 1

            errores.append(
                f'Fila {numero_fila}: el docente con cédula '
                f'{cedula} ya está registrado.'
            )

            continue

        Docente.objects.create(
            cedula=cedula,
            nombres=nombres,
            apellidos=apellidos,
            area=area,
            especialidad=especialidad,
            telefono=telefono,
            correo=correo,
            turno=turno,
            activo=convertir_activo(activo),
        )

        importados += 1

    return {
        'importados': importados,
        'duplicados': duplicados,
        'errores': errores,
        'cantidad_errores': len(errores),
    }