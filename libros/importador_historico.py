import re
import unicodedata
from datetime import date, datetime

from django.db import transaction
from openpyxl import load_workbook

from .models import Ejemplar, ImportacionLibros, Libro


def limpiar_texto(valor):
    if valor is None:
        return ''

    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)

    return str(valor).strip()


def normalizar_texto(valor):
    texto = limpiar_texto(valor).lower()
    texto = texto.strip('()[]{} ')

    texto = unicodedata.normalize(
        'NFD',
        texto,
    )

    texto = ''.join(
        caracter
        for caracter in texto
        if unicodedata.category(caracter) != 'Mn'
    )

    texto = re.sub(
        r'\s+',
        ' ',
        texto,
    )

    return texto.strip()


def normalizar_encabezado(valor):
    texto = normalizar_texto(valor)

    texto = re.sub(
        r'[^a-z0-9 ]',
        '',
        texto,
    )

    return texto.strip()


def obtener_numero_inventario(valor):
    texto = limpiar_texto(valor)

    if not texto:
        return None

    if isinstance(valor, int):
        return valor if valor > 0 else None

    if isinstance(valor, float) and valor.is_integer():
        numero = int(valor)
        return numero if numero > 0 else None

    encontrados = re.findall(
        r'\d+',
        texto,
    )

    if not encontrados:
        return None

    numero = int(encontrados[-1])

    return numero if numero > 0 else None


def convertir_fecha(valor):
    if valor in (None, ''):
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    texto = limpiar_texto(valor)

    texto = texto.strip()
    texto = texto.strip('()[]{} ')
    texto = texto.replace('(', '')
    texto = texto.replace(')', '')
    texto = texto.strip()

    formatos = [
        '%d/%m/%Y',
        '%d-%m-%Y',
        '%Y-%m-%d',
        '%d/%m/%y',
        '%d-%m-%y',
    ]

    for formato in formatos:
        try:
            return datetime.strptime(
                texto,
                formato,
            ).date()
        except ValueError:
            continue

    return None

def normalizar_adquisicion(valor):
    texto = normalizar_texto(valor)

    if not texto:
        return (
            Ejemplar.FormaAdquisicion.NO_ESPECIFICADA
        )

    if 'compra' in texto:
        return Ejemplar.FormaAdquisicion.COMPRA

    if 'donacion' in texto:
        return Ejemplar.FormaAdquisicion.DONACION

    if 'transferencia' in texto:
        return Ejemplar.FormaAdquisicion.TRANSFERENCIA

    return Ejemplar.FormaAdquisicion.OTRO


def separar_ubicacion(valor):
    texto = limpiar_texto(valor).upper()

    if not texto:
        return '', ''

    texto = texto.replace('ESTANTERÍA', '')
    texto = texto.replace('ESTANTERIA', '')
    texto = texto.replace('ESTANTE', '')
    texto = texto.replace('BALDA', '')
    texto = texto.strip(' :-')

    partes = [
        parte.strip()
        for parte in texto.split('-')
        if parte.strip()
    ]

    if len(partes) >= 2:
        return partes[0], partes[1]

    # El Excel histórico normalmente contiene solamente
    # el número de la balda.
    return '', texto


def buscar_encabezados(hoja):
    equivalencias = {
        'codigo_acceso': {
            'codigo acceso',
            'codigo de acceso',
            'nro acceso',
            'numero acceso',
        },
        'fecha_acceso': {
            'fecha acceso',
            'fecha de acceso',
            'fecha ingreso',
            'fecha de ingreso',
        },
        'autor': {
            'autor',
            'autores',
        },
        'titulo': {
            'titulo',
            'titulo del libro',
            'descripcion del libro',
        },
        'proveedor': {
            'proveedor',
        },
        'procedencia': {
            'procedencia',
            'forma adquisicion',
            'forma de adquisicion',
        },
        'balda': {
            'balda',
            'ubicacion',
            'estante',
        },
        'observacion': {
            'observacion',
            'observaciones',
        },
    }

    maximo_filas = min(
        hoja.max_row,
        40,
    )

    for numero_fila in range(1, maximo_filas + 1):
        columnas = {}

        for numero_columna in range(
            1,
            hoja.max_column + 1,
        ):
            encabezado = normalizar_encabezado(
                hoja.cell(
                    row=numero_fila,
                    column=numero_columna,
                ).value
            )

            if not encabezado:
                continue

            if encabezado.startswith(
                (
                    'codigo acceso',
                    'codigo de acceso',
                    'nro acceso',
                    'numero acceso',
                )
            ):
                columnas['codigo_acceso'] = numero_columna

            elif encabezado.startswith(
                (
                    'fecha acceso',
                    'fecha de acceso',
                    'fecha ingreso',
                    'fecha de ingreso',
                )
            ):
                columnas['fecha_acceso'] = numero_columna

            elif encabezado.startswith(
                (
                    'autor',
                    'autores',
                )
            ):
                columnas['autor'] = numero_columna

            elif encabezado.startswith(
                (
                    'titulo',
                    'descripcion del libro',
                )
            ):
                columnas['titulo'] = numero_columna

            elif encabezado.startswith('proveedor'):
                columnas['proveedor'] = numero_columna

            elif encabezado.startswith(
                (
                    'procedencia',
                    'forma adquisicion',
                    'forma de adquisicion',
                )
            ):
                columnas['procedencia'] = numero_columna

            elif encabezado.startswith(
                (
                    'balda',
                    'ubicacion',
                    'estante',
                )
            ):
                columnas['balda'] = numero_columna

            elif encabezado.startswith(
                (
                    'observacion',
                    'observaciones',
                )
            ):
                columnas['observacion'] = numero_columna

        if (
            'codigo_acceso' in columnas
            and 'titulo' in columnas
            and 'autor' in columnas
        ):
            return numero_fila, columnas

    return None, {}


def valor_columna(hoja, fila, columnas, campo):
    numero_columna = columnas.get(campo)

    if not numero_columna:
        return None

    return hoja.cell(
        row=fila,
        column=numero_columna,
    ).value


def extraer_registros(ruta_archivo):
    libro_excel = load_workbook(
        ruta_archivo,
        read_only=True,
        data_only=True,
    )

    registros = []
    errores = []
    hojas_omitidas = []

    for hoja in libro_excel.worksheets:
        fila_encabezado, columnas = buscar_encabezados(
            hoja
        )

        if fila_encabezado is None:
            hojas_omitidas.append(hoja.title)
            continue

        categoria = hoja.title.strip()

        for numero_fila in range(
            fila_encabezado + 1,
            hoja.max_row + 1,
        ):
            titulo = limpiar_texto(
                valor_columna(
                    hoja,
                    numero_fila,
                    columnas,
                    'titulo',
                )
            )

            autor = limpiar_texto(
                valor_columna(
                    hoja,
                    numero_fila,
                    columnas,
                    'autor',
                )
            )

            codigo_original = limpiar_texto(
                valor_columna(
                    hoja,
                    numero_fila,
                    columnas,
                    'codigo_acceso',
                )
            )

            if not titulo:
                continue

            if normalizar_texto(titulo) in {
                'descripcion del libro',
                'titulo',
            }:
                continue

            proveedor = limpiar_texto(
                valor_columna(
                    hoja,
                    numero_fila,
                    columnas,
                    'proveedor',
                )
            )

            procedencia = limpiar_texto(
                valor_columna(
                    hoja,
                    numero_fila,
                    columnas,
                    'procedencia',
                )
            )

            balda_original = limpiar_texto(
                valor_columna(
                    hoja,
                    numero_fila,
                    columnas,
                    'balda',
                )
            )

            observaciones = limpiar_texto(
                valor_columna(
                    hoja,
                    numero_fila,
                    columnas,
                    'observacion',
                )
            )

            fecha_original = valor_columna(
                hoja,
                numero_fila,
                columnas,
                'fecha_acceso',
            )

            fecha_adquisicion = convertir_fecha(
                fecha_original
            )

            if (
                fecha_original not in (None, '')
                and fecha_adquisicion is None
            ):
                errores.append(
                    (
                        f'{hoja.title}, fila {numero_fila}: '
                        'fecha no reconocida; se importará '
                        'sin fecha.'
                    )
                )

            estanteria, balda = separar_ubicacion(
                balda_original
            )

            adquisicion = normalizar_adquisicion(
                procedencia
            )

            if (
                procedencia
                and adquisicion
                == Ejemplar.FormaAdquisicion.OTRO
            ):
                texto_procedencia = (
                    f'Procedencia original: {procedencia}'
                )

                if observaciones:
                    observaciones = (
                        f'{observaciones}\n'
                        f'{texto_procedencia}'
                    )
                else:
                    observaciones = texto_procedencia

            registros.append(
                {
                    'hoja': hoja.title,
                    'fila': numero_fila,
                    'categoria': categoria,
                    'codigo_original': codigo_original,
                    'numero_original': (
                        obtener_numero_inventario(
                            codigo_original
                        )
                    ),
                    'fecha_adquisicion': (
                        fecha_adquisicion
                    ),
                    'autor': autor,
                    'titulo': titulo,
                    'proveedor': proveedor,
                    'forma_adquisicion': adquisicion,
                    'estanteria': estanteria,
                    'balda': balda,
                    'observaciones': observaciones,
                }
            )

    libro_excel.close()

    return registros, errores, hojas_omitidas


def analizar_archivo(ruta_archivo):
    registros, errores, hojas_omitidas = (
        extraer_registros(ruta_archivo)
    )

    codigos = [
        registro['numero_original']
        for registro in registros
        if registro['numero_original'] is not None
    ]

    codigos_unicos = set(codigos)

    codigos_repetidos = (
        len(codigos) - len(codigos_unicos)
    )

    titulos = {
        (
            normalizar_texto(registro['titulo']),
            normalizar_texto(registro['autor']),
        )
        for registro in registros
    }

    sin_ubicacion = sum(
        1
        for registro in registros
        if not registro['estanteria']
        and not registro['balda']
    )

    sin_codigo = sum(
        1
        for registro in registros
        if registro['numero_original'] is None
    )

    return {
        'registros': len(registros),
        'titulos_estimados': len(titulos),
        'ejemplares_estimados': len(registros),
        'codigos_repetidos': codigos_repetidos,
        'sin_codigo': sin_codigo,
        'sin_ubicacion': sin_ubicacion,
        'errores': errores,
        'cantidad_errores': len(errores),
        'hojas_omitidas': hojas_omitidas,
    }


def siguiente_numero_para_importacion(
    numeros_bloqueados,
):
    if not numeros_bloqueados:
        return 1

    menor = min(numeros_bloqueados)
    mayor = max(numeros_bloqueados)

    for numero in range(menor, mayor + 1):
        if numero not in numeros_bloqueados:
            return numero

    return mayor + 1


@transaction.atomic
def importar_archivo(
    ruta_archivo,
    nombre_archivo,
    huella_archivo,
):
    if ImportacionLibros.objects.filter(
        huella_archivo=huella_archivo,
    ).exists():
        raise ValueError(
            'Este archivo ya fue importado anteriormente.'
        )

    registros, errores, hojas_omitidas = (
        extraer_registros(ruta_archivo)
    )

    if not registros:
        raise ValueError(
            'No se encontraron registros válidos.'
        )

    numeros_existentes = set(
        Ejemplar.objects.exclude(
            numero_inventario__isnull=True,
        ).values_list(
            'numero_inventario',
            flat=True,
        )
    )

    numeros_reservados = {
        registro['numero_original']
        for registro in registros
        if registro['numero_original'] is not None
    }

    numeros_bloqueados = (
        numeros_existentes | numeros_reservados
    )

    numeros_asignados = set(numeros_existentes)
    codigos_reasignados = 0
    sin_ubicacion = 0
    titulos_creados = 0
    ejemplares_creados = 0

    libros_existentes = {}

    for libro in Libro.objects.all():
        clave = (
            normalizar_texto(libro.titulo),
            normalizar_texto(libro.autor),
        )

        libros_existentes[clave] = libro

    codigos_originales_utilizados = set()

    for registro in registros:
        clave_libro = (
            normalizar_texto(registro['titulo']),
            normalizar_texto(registro['autor']),
        )

        libro = libros_existentes.get(clave_libro)

        if libro is None:
            libro = Libro.objects.create(
                titulo=registro['titulo'],
                autor=registro['autor'],
                categoria=registro['categoria'],
                isbn='',
                activo=True,
            )

            libros_existentes[clave_libro] = libro
            titulos_creados += 1

        numero_original = registro['numero_original']
        numero_asignado = numero_original

        codigo_repetido = (
            numero_original
            in codigos_originales_utilizados
        )

        codigo_ocupado = (
            numero_original in numeros_existentes
        )

        if (
            numero_original is None
            or codigo_repetido
            or codigo_ocupado
        ):
            numero_asignado = (
                siguiente_numero_para_importacion(
                    numeros_bloqueados
                    | numeros_asignados
                )
            )

            codigos_reasignados += 1

        if numero_original is not None:
            codigos_originales_utilizados.add(
                numero_original
            )

        numeros_asignados.add(numero_asignado)
        numeros_bloqueados.add(numero_asignado)

        if (
            not registro['estanteria']
            and not registro['balda']
        ):
            sin_ubicacion += 1

        Ejemplar.objects.create(
            libro=libro,
            numero_inventario=numero_asignado,
            codigo_anterior=registro[
                'codigo_original'
            ],
            estanteria=registro['estanteria'],
            balda=registro['balda'],
            proveedor=registro['proveedor'],
            estado=Ejemplar.Estado.DISPONIBLE,
            condicion=Ejemplar.Condicion.BUENO,
            forma_adquisicion=registro[
                'forma_adquisicion'
            ],
            fecha_adquisicion=registro[
                'fecha_adquisicion'
            ],
            observaciones=registro[
                'observaciones'
            ],
        )

        ejemplares_creados += 1

    importacion = ImportacionLibros.objects.create(
        nombre_archivo=nombre_archivo,
        huella_archivo=huella_archivo,
        titulos_creados=titulos_creados,
        ejemplares_creados=ejemplares_creados,
        codigos_reasignados=codigos_reasignados,
        registros_sin_ubicacion=sin_ubicacion,
    )

    return {
        'importacion': importacion,
        'titulos_creados': titulos_creados,
        'ejemplares_creados': ejemplares_creados,
        'codigos_reasignados': codigos_reasignados,
        'sin_ubicacion': sin_ubicacion,
        'errores': errores,
        'hojas_omitidas': hojas_omitidas,
    }